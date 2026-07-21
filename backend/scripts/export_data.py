from __future__ import annotations

import json
import logging
import logging.config
from datetime import datetime
from pathlib import Path
from typing import Any, Final

import pandas as pd

from backend.config.config import (
    BUSINESS_INSIGHTS_CSV,
    EXECUTIVE_SUMMARY_CSV,
    EXPORTS_DIR,
    FINANCIAL_SUMMARY_CSV,
    FINAL_DASHBOARD_DATA_CSV,
    LOGGER_NAME,
    PROJECT_NAME,
    build_production_logging_config,
)



_REQUIRED_COLUMNS_BY_DATASET: Final[dict[str, list[str]]] = {
    "dashboard": [
        "Shipment ID",
        "Product Category",
        "Product Name",
        "Vehicle Type",
        "Transport Mode",
        "Origin City",
        "Destination City",
        "Temperature",
        "Humidity",
        "Battery",
        "Delay Hours",
        "Risk Score",
    ],
    "executive": ["Stakeholder", "Executive Summary"],
    "business_insights": ["Insight", "Value"],
    "financial": ["Metric", "Value"],
}


def setup_logging() -> logging.Logger:
    """Configure production logging and return the project logger."""

    logging.config.dictConfig(build_production_logging_config())
    return logging.getLogger(LOGGER_NAME)


def load_dashboard_dataset(logger: logging.Logger) -> pd.DataFrame:
    """Load the final dashboard dataset used for Power BI."""

    if not FINAL_DASHBOARD_DATA_CSV.exists():
        raise FileNotFoundError(
            f"Final dashboard dataset not found: {FINAL_DASHBOARD_DATA_CSV}"
        )

    logger.info("Loading Dashboard dataset from %s", str(FINAL_DASHBOARD_DATA_CSV))
    return pd.read_csv(FINAL_DASHBOARD_DATA_CSV)


def load_executive_summary(logger: logging.Logger) -> pd.DataFrame:
    """Load the executive summary dataset."""

    if not EXECUTIVE_SUMMARY_CSV.exists():
        raise FileNotFoundError(
            f"Executive summary CSV not found: {EXECUTIVE_SUMMARY_CSV}"
        )

    logger.info(
        "Loading Executive Summary from %s", str(EXECUTIVE_SUMMARY_CSV)
    )
    return pd.read_csv(EXECUTIVE_SUMMARY_CSV)


def load_business_insights(logger: logging.Logger) -> pd.DataFrame:
    """Load the business insights dataset."""

    if not BUSINESS_INSIGHTS_CSV.exists():
        raise FileNotFoundError(
            f"Business insights CSV not found: {BUSINESS_INSIGHTS_CSV}"
        )

    logger.info(
        "Loading Business Insights from %s", str(BUSINESS_INSIGHTS_CSV)
    )
    return pd.read_csv(BUSINESS_INSIGHTS_CSV)


def load_financial_summary(logger: logging.Logger) -> pd.DataFrame:
    """Load the financial summary dataset."""

    if not FINANCIAL_SUMMARY_CSV.exists():
        raise FileNotFoundError(
            f"Financial summary CSV not found: {FINANCIAL_SUMMARY_CSV}"
        )

    logger.info(
        "Loading Financial Summary from %s", str(FINANCIAL_SUMMARY_CSV)
    )
    return pd.read_csv(FINANCIAL_SUMMARY_CSV)


def validate_dataset(logger: logging.Logger, df: pd.DataFrame, dataset_key: str) -> None:
    """Validate dataset prior to export.

    Validation:
    - Must not be empty.
    - Required columns must exist.
    - Missing values are logged as warnings.

    Args:
        logger: Logger instance.
        df: Dataset dataframe.
        dataset_key: dashboard/executive/business_insights/financial.

    Raises:
        ValueError: If dataset is empty or required columns are missing.
    """

    if df.empty:
        raise ValueError(f"Dataset '{dataset_key}' is empty; refusing export")

    required_columns = _REQUIRED_COLUMNS_BY_DATASET.get(dataset_key, [])
    if required_columns:
        missing = [c for c in required_columns if c not in df.columns]
        if missing:
            raise ValueError(
                f"Dataset '{dataset_key}' is missing required columns: {missing}"
            )

    na_counts = df.isna().sum()
    missing_value_cols = na_counts[na_counts > 0]
    if not missing_value_cols.empty:
        top_missing = missing_value_cols.sort_values(ascending=False).head(10)
        logger.warning(
            "Dataset '%s' has missing values. Top columns: %s",
            dataset_key,
            top_missing.to_dict(),
        )


def export_csv_files(
    logger: logging.Logger,
    df_map: dict[str, pd.DataFrame],
    destination_dir: Path,
) -> list[dict[str, Any]]:
    """Export datasets as CSV into the destination directory."""

    destination_dir.mkdir(parents=True, exist_ok=True)

    export_specs: dict[str, str] = {
        "dashboard": "final_dashboard_data.csv",
        "executive": "executive_summary.csv",
        "business_insights": "business_insights.csv",
        "financial": "financial_summary.csv",
    }

    export_records: list[dict[str, Any]] = []

    for key, df in df_map.items():
        csv_name = export_specs.get(key)
        if not csv_name:
            continue

        out_path = destination_dir / csv_name
        logger.info("Exporting CSV (%d rows) -> %s", len(df), str(out_path))
        df.to_csv(out_path, index=False)

        export_records.append(
            {
                "dataset": key,
                "format": "csv",
                "path": str(out_path),
                "file_size_bytes": int(out_path.stat().st_size),
                "row_count": int(len(df)),
                "column_count": int(len(df.columns)),
            }
        )

    return export_records


def export_excel_files(
    logger: logging.Logger,
    df_map: dict[str, pd.DataFrame],
    destination_dir: Path,
) -> list[dict[str, Any]]:
    """Export datasets as Excel (.xlsx) into the destination directory."""

    destination_dir.mkdir(parents=True, exist_ok=True)

    export_specs: dict[str, str] = {
        "dashboard": "final_dashboard_data.xlsx",
        "executive": "executive_summary.xlsx",
        "business_insights": "business_insights.xlsx",
        "financial": "financial_summary.xlsx",
    }

    export_records: list[dict[str, Any]] = []

    # Local import to avoid failing the whole module import if openpyxl isn't available.
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise ImportError("openpyxl is required for Excel export") from exc

    for key, df in df_map.items():
        xlsx_name = export_specs.get(key)
        if not xlsx_name:
            continue

        out_path = destination_dir / xlsx_name
        logger.info("Exporting Excel (%d rows) -> %s", len(df), str(out_path))

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Header
        for col_idx, col_name in enumerate(df.columns.tolist(), start=1):
            ws.cell(row=1, column=col_idx, value=str(col_name))

        # Body
        for row_idx, row in enumerate(
            df.itertuples(index=False, name=None), start=2
        ):
            for col_idx, value in enumerate(row, start=1):
                if isinstance(value, (dict, list)):
                    ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=json.dumps(value, ensure_ascii=False),
                    )
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(out_path)

        export_records.append(
            {
                "dataset": key,
                "format": "xlsx",
                "path": str(out_path),
                "file_size_bytes": int(out_path.stat().st_size),
                "row_count": int(len(df)),
                "column_count": int(len(df.columns)),
            }
        )

    return export_records


def generate_export_summary(
    exported_records: list[dict[str, Any]],
    export_destination: str,
    export_start_dt: datetime,
    export_end_dt: datetime,
    overall_success: bool,
) -> dict[str, Any]:
    """Generate the exporter summary payload."""

    total_files = len(exported_records)
    row_counts = [int(r.get("row_count", 0)) for r in exported_records]
    col_counts = [int(r.get("column_count", 0)) for r in exported_records]
    file_sizes = [int(r.get("file_size_bytes", 0)) for r in exported_records]

    return {
        "project": PROJECT_NAME,
        "export_destination": export_destination,
        "export_start": export_start_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "export_end": export_end_dt.strftime("%Y-%m-%d %H:%M:%S"),
        "total_exported_files": int(total_files),
        "dataset_row_counts": row_counts,
        "dataset_column_counts": col_counts,
        "dataset_file_sizes_bytes": file_sizes,
        "export_status": "success" if overall_success else "partial_failure",
        "overall_export_success": bool(overall_success),
    }


def save_export_summary(
    logger: logging.Logger,
    summary: dict[str, Any],
    destination_dir: Path,
) -> Path:
    """Save export summary JSON."""

    destination_dir.mkdir(parents=True, exist_ok=True)
    summary_path = destination_dir / "export_summary.json"
    logger.info("Saving export summary -> %s", str(summary_path))

    summary_path.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    return summary_path


def export_pipeline() -> None:
    """Final exporter: load datasets, validate them, and export to delivery folders."""

    logger = setup_logging()
    export_start = datetime.now()

    destination_dir = (EXPORTS_DIR / "reports" / "output").resolve()

    df_map: dict[str, pd.DataFrame] = {}
    exported_records: list[dict[str, Any]] = []
    overall_success = True

    try:
        df_map["dashboard"] = load_dashboard_dataset(logger)
        df_map["executive"] = load_executive_summary(logger)
        df_map["business_insights"] = load_business_insights(logger)
        df_map["financial"] = load_financial_summary(logger)

        for key, df in df_map.items():
            validate_dataset(logger=logger, df=df, dataset_key=key)

        exported_records.extend(
            export_csv_files(
                logger=logger,
                df_map=df_map,
                destination_dir=destination_dir,
            )
        )

        exported_records.extend(
            export_excel_files(
                logger=logger,
                df_map=df_map,
                destination_dir=destination_dir,
            )
        )

    except Exception:
        overall_success = False
        logger.exception("Export pipeline failed")
        raise
    finally:
        export_end = datetime.now()
        summary = generate_export_summary(
            exported_records=exported_records,
            export_destination=str(destination_dir),
            export_start_dt=export_start,
            export_end_dt=export_end,
            overall_success=overall_success,
        )
        try:
            save_export_summary(
                logger=logger,
                summary=summary,
                destination_dir=destination_dir,
            )
        except Exception:
            logger.exception("Failed to save export summary")


def main() -> None:
    """CLI entry point."""

    export_pipeline()


if __name__ == "__main__":
    main()
